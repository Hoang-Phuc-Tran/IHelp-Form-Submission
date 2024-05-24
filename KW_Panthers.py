import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from datetime import datetime
from PIL import Image, ImageTk
from openpyxl import load_workbook
import re

# Load Excel workbook
wb = load_workbook("basketball.xlsx")
ws = wb.active
next_row = ws.max_row + 1

repeat = False

def yes_clicked():
    global repeat
    repeat = True
    root.deiconify()
    submission_window.destroy()

def no_clicked():
    global repeat
    repeat = False
    submission_window.destroy()
    root.destroy()

def submission_message():
    global submission_window
    submission_window = tk.Toplevel()
    submission_window.config(bg="white")
    submission_window.title("Submission successful")

    thankyouLabel = tk.Label(submission_window, text="Thank You!", font=("Helvetica", 34, "bold"), fg="purple")
    thankyouLabel.config(bg="white")
    thankyouLabel.pack(pady=(30, 20))

    receivedLabel = tk.Label(submission_window, text="Your submission has been received.", font=("Helvetica", 14))
    receivedLabel.config(bg="white")
    receivedLabel.pack()

    global recnew

    rec = Image.open("received.jpg")
    recrs = rec.resize((350, 250))
    recnew = ImageTk.PhotoImage(recrs)
    recLabel = tk.Label(submission_window, image=recnew)
    recLabel.config(bg="white")
    recLabel.pack()

    againLabel = tk.Label(submission_window, text="Would you like to add another submission?", font=("Helvetica", 14, "bold"), bg="white")
    againLabel.pack()

    yesButton = tk.Button(submission_window, text="Yes", width=8, height=1, font=("Helvetica", 14, "bold"), command=yes_clicked, bg="#b16df4", fg="white")
    yesButton.bind("<Enter>", lambda event: yesButton.config(bg="#ffe85a", fg="purple"))
    yesButton.bind("<Leave>", lambda event: yesButton.config(bg="#b16df4", fg="white"))
    yesButton.pack(side=LEFT, padx=(115, 0))

    noButton = tk.Button(submission_window, text="No", width=8, height=1, font=("Helvetica", 14, "bold"), command=no_clicked, bg="#b16df4", fg="white")
    noButton.pack(side=RIGHT, padx=(0, 115))
    noButton.bind("<Enter>", lambda event: noButton.config(bg="#ffe85a", fg="purple"))
    noButton.bind("<Leave>", lambda event: noButton.config(bg="#b16df4", fg="white"))

    width = 500
    height = 550

    x = (submission_window.winfo_screenwidth() - width) // 2
    y = (submission_window.winfo_screenheight() - height) // 2

    submission_window.geometry(f'{width}x{height}+{x}+{y}')

def clickIn(event, entry):
    text = entry.get()
    entry.config(fg="black")
    if text == placeholders[entry]:
        entry.delete(0, tk.END)

def clickOut(event, entry):
    text = entry.get()
    if text == "":
        entry.insert(0, placeholders[entry])
        entry.config(fg="grey")

# Function to validate names
def validate_name(entry, label, error_label):
    if not entry.get().replace(" ", "").isalpha():
        label.config(fg="red")
        error_label.config(text="Please Enter a valid name", fg="red")
        error_label.grid(row=label.grid_info()['row'], column=2, sticky="W")
        return False
    else:
        label.config(fg="black")
        error_label.config(text="")
        return True

def calculateage():
    try:
        dob = datetime.strptime(f"{yearEntry.get()}-{monthEntry.get()}-{dayEntry.get()}", "%Y-%m-%d")
        today = datetime.today()
        ageYear = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        ageMonth = today.month - dob.month if today.month >= dob.month else 12 - (dob.month - today.month)
        ageDisplay.config(text=f"{ageYear} Years and {ageMonth} months")
        dateOfBirthLabel.config(fg="black")
        return True
    except:
        dateOfBirthLabel.config(fg="red")
        return False

def check_phone_number_pattern(entry):
    patterns = [r'^\d{9}$', r'^\d{10}$', r'^\d{11}$', r'^\d{12}$', r'^\+\d{9}$',
                r'^\+\d{10}$', r'^\+\d{11}$', r'^\+\d{12}$', r'^\d{3} \d{3} \d{4}$',
                r'^\d{3}-\d{3}-\d{4}$']
    for pattern in patterns:
        if re.match(pattern, entry.get()):
            return True
    return False

def validate_phone_number():
    try:
        if (phoneNumberEntry.get() == placeholders[phoneNumberEntry] or not check_phone_number_pattern(phoneNumberEntry)):
            phoneNumberLabel.config(fg="red")
            phoneNumberError.config(text="Please enter a valid phone number", fg="red")
            phoneNumberError.grid(row=7, column=2, sticky="W")
            return False
        else:
            phoneNumberError.config(text="")
            phoneNumberLabel.config(fg="black")
            return True
    except:
        phoneNumberLabel.config(fg="red")
        phoneNumberError.config(text="Please enter a valid phone number", fg="red")
        phoneNumberError.grid(row=7, column=2, sticky="W")
        return False

def validate_email():
    stringvalidation = "@" not in emailEntry.get() or "." not in emailEntry.get()
    if stringvalidation or emailEntry.get() == placeholders[emailEntry]:
        emailLabel.config(fg="red")
        emailError.config(text="Please enter a valid email", fg="red")
        emailError.grid(row=8, column=2, sticky="W")
        return False
    else:
        emailError.config(text="")
        emailLabel.config(fg="black")
        return True

def validate_parent_name():
    valid = True
    if not parentFirstNameEntry.get().replace(" ", "").isalpha():
        parentFirstNameLabel.config(fg="red")
        parentFirstNameError.config(text="Please enter a valid name", fg="red")
        parentFirstNameError.grid(row=9, column=2, sticky="W")
        valid = False
    else:
        parentFirstNameLabel.config(fg="black")
        parentFirstNameError.config(text="")

    if parentMiddleNameEntry.get() and not parentMiddleNameEntry.get().replace(" ", "").isalpha():
        parentMiddleNameLabel.config(fg="red")
        parentMiddleNameError.config(text="Please enter a valid name", fg="red")
        parentMiddleNameError.grid(row=10, column=2, sticky="W")
        valid = False
    else:
        parentMiddleNameLabel.config(fg="black")
        parentMiddleNameError.config(text="")

    if not parentLastNameEntry.get().replace(" ", "").isalpha():
        parentLastNameLabel.config(fg="red")
        parentLastNameError.config(text="Please enter a valid name", fg="red")
        parentLastNameError.grid(row=11, column=2, sticky="W")
        valid = False
    else:
        parentLastNameLabel.config(fg="black")
        parentLastNameError.config(text="")

    return valid

def validate_parent_phone_number():
    try:
        if (not check_phone_number_pattern(parentPhoneNumberEntry)):
            parentPhoneNumberLabel.config(fg="red")
            parentPhoneNumberError.config(text="Please enter a valid phone number", fg="red")
            parentPhoneNumberError.grid(row=12, column=2, sticky="W")
            return False
        else:
            parentPhoneNumberError.config(text="")
            parentPhoneNumberLabel.config(fg="black")
            return True
    except:
        parentPhoneNumberLabel.config(fg="red")
        parentPhoneNumberError.config(text="Please enter a valid phone number", fg="red")
        parentPhoneNumberError.grid(row=12, column=2, sticky="W")
        return False

def validate_Street():
    if len(StreetEntry.get().replace(" ", "")) == 0:
        StreetLabel.config(fg="red")
        StreetError.config(text="Please enter a valid Street Name", fg="red")
        StreetError.grid(row=13, column=2, sticky="W")
        return False
    else:
        StreetError.config(text="")
        StreetLabel.config(fg="black")
        return True

def validate_city():
    if not cityEntry.get().replace(" ", "").isalpha():
        cityLabel.config(fg="red")
        cityError.config(text="Please enter a City", fg="red")
        cityError.grid(row=14, column=2, sticky="W")
        return False
    else:
        cityError.config(text="")
        cityLabel.config(fg="black")
        return True

def validate_PostalCode():
    pattern = r'^[a-zA-Z]\d[a-zA-Z] \d[a-zA-Z]\d$'
    if len(PostalCodeEntry.get().replace(" ", "")) == 0 or not re.match(pattern, PostalCodeEntry.get()):
        PostalCodeLabel.config(fg="red")
        PostalCodeError.config(text="Please enter a valid Postal Code", fg="red")
        PostalCodeError.grid(row=15, column=2, sticky="W")
        return False
    else:
        PostalCodeError.config(text="")
        PostalCodeLabel.config(fg="black")
        return True

def submit():
    global next_row
    first_name_valid = validate_name(firstNameEntry, firstNameLabel, firstNameError)
    last_name_valid = validate_name(lastNameEntry, lastNameLabel, lastNameError)
    parent_name_valid = validate_parent_name()

    if first_name_valid and last_name_valid and calculateage() and validate_phone_number() and validate_email() and parent_name_valid and validate_parent_phone_number() and validate_Street() and validate_city() and validate_PostalCode():

        SubmissionDate = datetime.today().date()
        firstName = firstNameEntry.get()
        middleName = middleNameEntry.get()
        lastName = lastNameEntry.get()
        fullName = f"{firstName} {middleName} {lastName}".strip()

        DOB = datetime.strptime(f"{yearEntry.get()}-{monthEntry.get()}-{dayEntry.get()}", "%Y-%m-%d").date()
        today = datetime.today()
        Age = today.year - DOB.year - ((today.month, today.day) < (DOB.month, DOB.day))
        Gender = genderValue.get()
        PhNumber = phoneNumberEntry.get()
        Email = emailEntry.get()
        ParentFirstName = parentFirstNameEntry.get()
        ParentMiddleName = parentMiddleNameEntry.get()
        ParentLastName = parentLastNameEntry.get()
        ParentFullName = f"{ParentFirstName} {ParentMiddleName} {ParentLastName}".strip()
        Address = f"{StreetEntry.get()}, {cityEntry.get()}, {PostalCodeEntry.get()}"
        ParentPhNumber = parentPhoneNumberEntry.get()
        Remarks = remarksEntry.get("1.0", "end-1c")

        ws[f"A{next_row}"] = SubmissionDate
        ws[f"B{next_row}"] = fullName
        ws[f"C{next_row}"] = DOB
        ws[f"D{next_row}"] = Age
        ws[f"E{next_row}"] = Gender
        ws[f"F{next_row}"] = PhNumber
        ws[f"G{next_row}"] = Email
        ws[f"H{next_row}"] = ParentFullName
        ws[f"I{next_row}"] = Address
        ws[f"J{next_row}"] = ParentPhNumber
        ws[f"K{next_row}"] = Remarks

        wb.save("basketball.xlsx")
        next_row += 1

        firstNameEntry.delete(0, tk.END)
        middleNameEntry.delete(0, tk.END)
        lastNameEntry.delete(0, tk.END)
        yearEntry.delete(0, tk.END)
        monthEntry.delete(0, tk.END)
        dayEntry.delete(0, tk.END)
        phoneNumberEntry.delete(0, tk.END)
        emailEntry.delete(0, tk.END)
        parentFirstNameEntry.delete(0, tk.END)
        parentMiddleNameEntry.delete(0, tk.END)
        parentLastNameEntry.delete(0, tk.END)
        StreetEntry.delete(0, tk.END)
        cityEntry.delete(0, tk.END)
        PostalCodeEntry.delete(0, tk.END)
        parentPhoneNumberEntry.delete(0, tk.END)
        remarksEntry.delete("1.0", tk.END)
        root.withdraw()

        submission_message()

def onmousescroll(event):
    my_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

root = tk.Tk()
root.title("Basketball Members")
root.geometry("900x900")
root.resizable(0, 1)

main_frame = Frame(root)
main_frame.pack(fill=BOTH, expand=1)

my_canvas = Canvas(main_frame)
my_canvas.pack(side=LEFT, fill=BOTH, expand=1)

my_scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
my_scrollbar.pack(side=RIGHT, fill=Y)

my_canvas.configure(yscrollcommand=my_scrollbar.set)
my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all")))
my_canvas.bind_all("<MouseWheel>", onmousescroll)

content_frame = Frame(my_canvas)

my_canvas.create_window((0, 0), window=content_frame, anchor="nw")

ball = Image.open('basketball_PNG102482.png')
ballrs = ball.resize((180, 180))
ballnew = ImageTk.PhotoImage(ballrs)
ballLabel = tk.Label(content_frame, image=ballnew)
ballLabel.grid(row=0, column=1, pady="5px")

iHelp = Image.open('ihelp Logo PNG (Facebook 180pix).png')
iHelprs = iHelp.resize((103, 50))
iHelpnew = ImageTk.PhotoImage(iHelprs)
iHelpLabel = tk.Label(content_frame, image=iHelpnew)
iHelpLabel.grid(row=0, column=2, sticky="NE", pady=35, padx=(50, 35))

titleLabel = tk.Label(content_frame, text="BASKETBALL", font=("Arial", 25, "bold"), fg="#e93514")
titleLabel.grid(row=0, column=0, sticky="ne", padx=(30, 0), pady=30)

subtitleLabel = tk.Label(content_frame, text="KW PANTHERS", font=("Arial", 13, "bold"))
subtitleLabel.grid(row=0, column=0, rowspan=2, sticky="ne", pady=(75, 0), padx=(0, 55))

# First Name
firstNameLabel = tk.Label(content_frame, text='First Name: ', font=("Arial", 11))
firstNameLabel.grid(row=1, column=0, sticky="W", padx="15px")
firstNameEntry = tk.Entry(content_frame, font=("Arial", 11))
firstNameEntry.grid(row=1, column=1, sticky="W", pady="10px", padx="5px")
firstNameEntry.bind("<FocusOut>", lambda event: validate_name(firstNameEntry, firstNameLabel, firstNameError))
firstNameError = tk.Label(content_frame)

# Middle Name
middleNameLabel = tk.Label(content_frame, text='Middle Name (optional): ', font=("Arial", 11))
middleNameLabel.grid(row=2, column=0, sticky="W", padx="15px")
middleNameEntry = tk.Entry(content_frame, font=("Arial", 11))
middleNameEntry.grid(row=2, column=1, sticky="W", pady="10px", padx="5px")

# Last Name
lastNameLabel = tk.Label(content_frame, text='Last Name: ', font=("Arial", 11))
lastNameLabel.grid(row=3, column=0, sticky="W", padx="15px")
lastNameEntry = tk.Entry(content_frame, font=("Arial", 11))
lastNameEntry.grid(row=3, column=1, sticky="W", pady="10px", padx="5px")
lastNameEntry.bind("<FocusOut>", lambda event: validate_name(lastNameEntry, lastNameLabel, lastNameError))
lastNameError = tk.Label(content_frame)

# Date of Birth
dateOfBirthLabel = tk.Label(content_frame, text="Date of Birth: ", font=("Arial", 11))
dateOfBirthLabel.grid(row=4, column=0, sticky="W", padx="15px")

validyear = False
validmonth = False
validdate = False

def yearEntryValidation(event):
    text = yearEntry.get()
    global validyear
    validyear = False
    validtext = ""
    for c in text:
        if c.isdigit():
            validtext += c
    validtext = validtext[0:4]
    yearEntry.delete(0, tk.END)
    yearEntry.insert(0, validtext)
    try:
        if int(validtext) < 1950 or int(validtext) > 2024:
            yearEntry.config(fg="red")
            validyear = False
        else:
            yearEntry.config(fg="black")
            validyear = True
    except:
        pass
    if validyear and validmonth and validdate:
        calculateage()

def monthEntryValidation(event):
    text = monthEntry.get()
    global validmonth
    validmonth = False
    validtext = ""
    for c in text:
        if c.isdigit():
            validtext += c
    validtext = validtext[0:2]
    monthEntry.delete(0, tk.END)
    monthEntry.insert(0, validtext)
    if dayEntry.get() != "DD":
        validate_date()
    try:
        if int(validtext) < 1 or int(validtext) > 12:
            monthEntry.config(fg="red")
            validmonth = False
        else:
            monthEntry.config(fg="black")
            validmonth = True
    except:
        pass
    if len(monthEntry.get()) == 2 and validmonth:
        yearEntry.focus_set()

    if validyear and validmonth and validdate:
        calculateage()

def dayEntryValidation(event):
    text = dayEntry.get()
    validtext = ""
    for c in text:
        if c.isdigit():
            validtext += c
    validtext = validtext[0:2]
    dayEntry.delete(0, tk.END)
    dayEntry.insert(0, validtext)
    validate_date()
    if len(dayEntry.get()) == 2 and validdate:
        monthEntry.focus_set()

def validate_date():
    global validdate
    validdate = False
    try:
        match monthEntry.get():
            case "01" | "03" | "05" | "07" | "08" | "10" | "12":
                if int(dayEntry.get()) < 1 or int(dayEntry.get()) > 31:
                    dayEntry.config(fg="red")
                    validdate = False
                else:
                    dayEntry.config(fg="black")
                    validdate = True
            case "04" | "06" | "09" | "11":
                if int(dayEntry.get()) < 1 or int(dayEntry.get()) > 30:
                    dayEntry.config(fg="red")
                    validdate = False
                else:
                    dayEntry.config(fg="black")
                    validdate = True
            case "02":
                if int(dayEntry.get()) < 1 or int(dayEntry.get()) > 28:
                    dayEntry.config(fg="red")
                    validdate = False
                else:
                    dayEntry.config(fg="black")
                    validdate = True
            case _:
                if int(dayEntry.get()) < 1 or int(dayEntry.get()) > 31:
                    dayEntry.config(fg="red")
                    validdate = False
                else:
                    dayEntry.config(fg="black")
                    validdate = True
    except:
        pass

    if validyear and validmonth and validdate:
        calculateage()

dayLabel = tk.Label(content_frame, text="DD", font=("Arial", 10, "bold"))
dayLabel.grid(row=4, column=1, sticky="NW", padx="5px", pady=(0, 40))
dayEntry = tk.Entry(content_frame, font=("Arial", 11), width=3)
dayEntry.config(fg="grey")
dayEntry.insert(0, "DD")
dayEntry.grid(row=4, column=1, sticky="W", pady="10px", padx="5px")
dayEntry.bind('<KeyRelease>', dayEntryValidation)
dayEntry.bind("<FocusIn>", lambda event: clickIn(event, dayEntry))
dayEntry.bind("<FocusOut>", lambda event: clickOut(event, dayEntry))

hyphen1_label = tk.Label(content_frame, text="-", font=("Arial", 11, "bold"))
hyphen1_label.grid(row=4, column=1, sticky="W", pady="10px", padx="30px")

monthLabel = tk.Label(content_frame, text="MM", font=("Arial", 10, "bold"))
monthLabel.grid(row=4, column=1, sticky="NW", padx="45px", pady=(0, 40))
monthEntry = tk.Entry(content_frame, font=("Arial", 11), width=3)
monthEntry.config(fg="grey")
monthEntry.insert(0, "MM")
monthEntry.grid(row=4, column=1, sticky="W", pady="10px", padx="45px")
monthEntry.bind('<KeyRelease>', monthEntryValidation)
monthEntry.bind("<FocusIn>", lambda event: clickIn(event, monthEntry))
monthEntry.bind("<FocusOut>", lambda event: clickOut(event, monthEntry))

hyphen2_label = tk.Label(content_frame, text="-", font=("Arial", 11, "bold"))
hyphen2_label.grid(row=4, column=1, sticky="W", pady="10px", padx="72px")

yearLabel = tk.Label(content_frame, text="YYYY", font=("Arial", 10, "bold"))
yearLabel.grid(row=4, column=1, sticky="NW", padx="85px", pady=(0, 40))
yearEntry = tk.Entry(content_frame, font=("Arial", 11), width=5)
yearEntry.config(fg="grey")
yearEntry.insert(0, "YYYY")
yearEntry.grid(row=4, column=1, sticky="W", pady="10px", padx="85px")
yearEntry.bind('<KeyRelease>', yearEntryValidation)
yearEntry.bind("<FocusIn>", lambda event: clickIn(event, yearEntry))
yearEntry.bind("<FocusOut>", lambda event: clickOut(event, yearEntry))

# Age
ageLabel = tk.Label(content_frame, text="Age: ", font=("Arial", 11))
ageLabel.grid(row=5, column=0, sticky="W", padx="15px")
ageDisplay = tk.Label(content_frame, text="", font=("Arial", 11))
ageDisplay.grid(row=5, column=1, sticky="W", pady="10px", padx="5px")

# Gender
genderLabel = tk.Label(content_frame, text="Gender", font=("Arial", 11))
genderLabel.grid(row=6, column=0, sticky="W", padx="15px")
genderValue = StringVar(content_frame)
genderValue.set("Male")

maleRadio = Radiobutton(content_frame, text="Male", font=("Arial", 11), variable=genderValue, value="Male")
maleRadio.grid(row=6, column=1, sticky="W", pady="10px", padx=(0, 50))

femaleRadio = Radiobutton(content_frame, text="Female", font=("Arial", 11), variable=genderValue, value="Female")
femaleRadio.grid(row=6, column=1, pady="10px", sticky="W", padx=(90, 0))

otherRadio = Radiobutton(content_frame, text="Other", font=("Arial", 11), variable=genderValue, value="Other")
otherRadio.grid(row=6, column=1, sticky="W", pady="10px", padx=(200, 0))

# Phone Number
phoneNumberLabel = tk.Label(content_frame, text="Phone Number", font=("Arial", 11))
phoneNumberLabel.grid(row=7, column=0, sticky="W", padx="15px")
phoneNumberEntry = tk.Entry(content_frame, font=("Arial", 11))
phoneNumberEntry.config(fg="grey")
phoneplaceholder = "1234567890"
phoneNumberEntry.insert(0, phoneplaceholder)
phoneNumberEntry.bind("<FocusIn>", lambda event: clickIn(event, phoneNumberEntry))
phoneNumberEntry.bind("<FocusOut>", lambda event: clickOut(event, phoneNumberEntry))
phoneNumberEntry.grid(row=7, column=1, sticky="W", pady="10px", padx="5px")
phoneNumberEntry.config(validate="focusout", validatecommand=validate_phone_number)
phoneNumberError = tk.Label(content_frame)

# Email
emailLabel = tk.Label(content_frame, text="Email: ", font=("Arial", 11))
emailLabel.grid(row=8, column=0, sticky="W", padx="15px")
emailEntry = tk.Entry(content_frame, font=("Arial", 11))
emailEntry.config(fg="grey")
emailplaceholderText = "example@mail.com"
emailEntry.insert(0, emailplaceholderText)
emailEntry.bind("<FocusIn>", lambda event: clickIn(event, emailEntry))
emailEntry.bind("<FocusOut>", lambda event: clickOut(event, emailEntry))
emailEntry.grid(row=8, column=1, sticky="W", pady="10px", padx="5px")
emailEntry.config(validate="focusout", validatecommand=validate_email)
emailError = tk.Label(content_frame)

# Parent First Name
parentFirstNameLabel = tk.Label(content_frame, text="Parent First Name: ", font=("Arial", 11))
parentFirstNameLabel.grid(row=9, column=0, sticky="W", padx="15px")
parentFirstNameEntry = tk.Entry(content_frame, font=("Arial", 11))
parentFirstNameEntry.grid(row=9, column=1, sticky="W", pady="10px", padx="5px")
parentFirstNameEntry.bind("<FocusOut>", lambda event: validate_name(parentFirstNameEntry, parentFirstNameLabel, parentFirstNameError))
parentFirstNameError = tk.Label(content_frame)

# Parent Middle Name
parentMiddleNameLabel = tk.Label(content_frame, text="Parent Middle Name (optional): ", font=("Arial", 11))
parentMiddleNameLabel.grid(row=10, column=0, sticky="W", padx="15px")
parentMiddleNameEntry = tk.Entry(content_frame, font=("Arial", 11))
parentMiddleNameEntry.grid(row=10, column=1, sticky="W", pady="10px", padx="5px")
parentMiddleNameEntry.bind("<FocusOut>", lambda event: validate_name(parentMiddleNameEntry, parentMiddleNameLabel, parentMiddleNameError))
parentMiddleNameError = tk.Label(content_frame)

# Parent Last Name
parentLastNameLabel = tk.Label(content_frame, text="Parent Last Name: ", font=("Arial", 11))
parentLastNameLabel.grid(row=11, column=0, sticky="W", padx="15px")
parentLastNameEntry = tk.Entry(content_frame, font=("Arial", 11))
parentLastNameEntry.grid(row=11, column=1, sticky="W", pady="10px", padx="5px")
parentLastNameEntry.bind("<FocusOut>", lambda event: validate_name(parentLastNameEntry, parentLastNameLabel, parentLastNameError))
parentLastNameError = tk.Label(content_frame)

# Parent Phone Number
parentPhoneNumberLabel = tk.Label(content_frame, text="Parent Phone Number", font=("Arial", 11))
parentPhoneNumberLabel.grid(row=12, column=0, sticky="W", padx="15px")
parentPhoneNumberEntry = tk.Entry(content_frame, font=("Arial", 11))
parentPhoneNumberEntry.grid(row=12, column=1, sticky="W", pady="10px", padx="5px")
parentPhoneNumberEntry.config(validate="focusout", validatecommand=validate_parent_phone_number)
parentPhoneNumberError = tk.Label(content_frame)

# Address
StreetLabel = tk.Label(content_frame, text="Street", font=("Arial", 11))
StreetLabel.grid(row=13, column=0, sticky="W", padx="15px")
StreetEntry = tk.Entry(content_frame, font=("Arial", 11))
StreetEntry.grid(row=13, column=1, sticky="W", pady="10px", padx="5px")
StreetEntry.config(validate="focusout", validatecommand=validate_Street)
StreetError = tk.Label(content_frame)

cityLabel = tk.Label(content_frame, text="City", font=("Arial", 11))
cityLabel.grid(row=14, column=0, sticky="W", padx="15px")
cityEntry = tk.Entry(content_frame, font=("Arial", 11))
cityEntry.grid(row=14, column=1, sticky="W", pady="10px", padx="5px")
cityEntry.config(validate="focusout", validatecommand=validate_city)
cityError = tk.Label(content_frame)

PostalCodeLabel = tk.Label(content_frame, text="PostalCode", font=("Arial", 11))
PostalCodeLabel.grid(row=15, column=0, sticky="W", padx="15px")
PostalCodeEntry = tk.Entry(content_frame, font=("Arial", 11))
PostalCodeEntry.config(fg="grey")
PostalCodeplaceholder = "X1X 1X1"
PostalCodeEntry.insert(0, PostalCodeplaceholder)
PostalCodeEntry.bind("<FocusIn>", lambda event: clickIn(event, PostalCodeEntry))
PostalCodeEntry.bind("<FocusOut>", lambda event: clickOut(event, PostalCodeEntry))
PostalCodeEntry.grid(row=15, column=1, sticky="W", pady="10px", padx="5px")
PostalCodeEntry.config(validate="focusout", validatecommand=validate_PostalCode)
PostalCodeError = tk.Label(content_frame)

# Remarks
remarksLabel = tk.Label(content_frame, text="Remarks", font=("Arial", 11))
remarksLabel.grid(row=16, column=0, sticky="W", padx="15px")
remarksEntry = tk.Text(content_frame, height=6, width=50)
remarksEntry.grid(row=16, column=1, sticky="W", pady="10px", padx="5px")

# Submit Button
submitButton = tk.Button(content_frame, width=25, height=2, text="Submit", bg="#f95f30", font=("Arial", 13, "bold"), command=submit)
submitButton.grid(row=17, column=1, columnspan=1, pady=(0, 50))
submitButton.bind("<Enter>", lambda event: submitButton.config(bg="#a5fc03", fg="red"))
submitButton.bind("<Leave>", lambda event: submitButton.config(bg="#f95f30", fg="black"))

# Placeholders
placeholders = {
    yearEntry: "YYYY",
    monthEntry: "MM",
    dayEntry: "DD",
    emailEntry: "example@mail.com",
    phoneNumberEntry: "1234567890",
    PostalCodeEntry: "X1X 1X1"
}

root.mainloop()
